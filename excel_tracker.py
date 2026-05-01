"""
AI Job Agent - Excel Tracker
Appends each run's results to job_tracker.xlsx.
Columns mirror the full jobspy field set (18 columns).
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper import _format_salary

logger = logging.getLogger(__name__)

TRACKER_FILE  = "job_tracker.xlsx"
DATA_START    = 5     # first data row (rows 1-3 title/sub/spacer, row 4 headers)
NUM_COLS      = 18    # A → R

# ── Styles ────────────────────────────────────────────────────
TITLE_BG  = "0F3172"
HEADER_BG = "1A56DB"
ALT_BG    = "EEF3FB"
BCLR      = "C5CEDF"
thin      = Side(style="thin", color=BCLR)
CB        = Border(left=thin, right=thin, top=thin, bottom=thin)

STATUS_COLORS = {
    "To Apply":  ("FFF3CD", "856404"),
    "Applied ✓": ("D1FAE5", "065F46"),
    "Skipped":   ("F3F4F6", "6B7280"),
    "Interview": ("DBEAFE", "1E40AF"),
    "Rejected":  ("FEE2E2", "991B1B"),
    "Offer":     ("D1FAE5", "065F46"),
}
SITE_COLORS = {
    "LinkedIn":     ("E8F0FE", "1A56DB"),
    "Indeed":       ("FFF3E0", "E65100"),
    "ZipRecruiter": ("F3E8FF", "7C3AED"),
    "Glassdoor":    ("E8F5E9", "2E7D32"),
    "Google":       ("FCE4EC", "C62828"),
}

def _bf(bold=False, color="000000", size=9):
    return Font(name="Arial", size=size, bold=bold, color=color)

def _ca(h="center"):
    return Alignment(horizontal=h, vertical="center", wrap_text=False)

def _la():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _find_next_row(ws) -> int:
    """Find the first empty data row (checks col B = Date Run)."""
    row = DATA_START
    while ws.cell(row=row, column=2).value is not None:
        row += 1
    return row


def _style_row(ws, row: int, source_label: str, status: str, alt: bool):
    """Apply borders, alternating background, and badge colours to a row."""
    for ci in range(1, NUM_COLS + 1):
        c = ws.cell(row=row, column=ci)
        c.border = CB

        if ci == 3:    # Source badge
            bg, fg = SITE_COLORS.get(source_label, ("F3F4F6", "374151"))
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = _bf(bold=True, color=fg)
            c.alignment = _ca()

        elif ci == 11:  # Days Since Posted (formula)
            c.font = _bf(bold=True, color="1A56DB")
            c.fill = PatternFill("solid", fgColor="EEF3FB")
            c.alignment = _ca()

        elif ci == 17:  # Status badge
            bg, fg = STATUS_COLORS.get(status, ("FFFFFF", "000000"))
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = _bf(bold=True, color=fg)
            c.alignment = _ca()

        elif ci in (5, 13, 14, 15, 16, 18):  # wrap-text cols
            c.font = _bf(color="1A56DB" if ci in (13, 14) else "000000")
            c.alignment = _la()
            if alt and ci not in (3, 11, 17):
                c.fill = PatternFill("solid", fgColor=ALT_BG)

        else:
            c.font = _bf()
            c.alignment = _ca()
            if alt and ci not in (3, 11, 17):
                c.fill = PatternFill("solid", fgColor=ALT_BG)


def append_jobs_to_tracker(results: List[Dict], tracker_path: str = TRACKER_FILE) -> int:
    """
    Append job results to the Excel tracker.

    Each result dict should have:
        job: full job dict from scraper (all jobspy fields)
        resume_path: str — path to the saved .docx file

    Returns number of rows added.
    """
    if not Path(tracker_path).exists():
        logger.warning(f"Tracker not found at '{tracker_path}'. Re-run setup or check the file.")
        return 0

    wb = load_workbook(tracker_path)
    ws = wb["Job Tracker"]

    today_str = datetime.today().strftime("%Y-%m-%d")
    rows_added = 0

    for result in results:
        job         = result.get("job", {})
        resume_path = result.get("resume_path", "")
        resume_name = Path(resume_path).name if resume_path else ""

        # ── Source label ──────────────────────────────────────
        raw_source = job.get("source", "")
        source_label = {
            "linkedin":      "LinkedIn",
            "indeed":        "Indeed",
            "zip_recruiter": "ZipRecruiter",
            "glassdoor":     "Glassdoor",
            "google":        "Google",
        }.get(raw_source.lower(), raw_source.title() or "—")

        next_row = _find_next_row(ws)
        alt      = (next_row - DATA_START) % 2 == 1
        row_num  = next_row - DATA_START + 1

        # ── Write 18 columns ─────────────────────────────────
        # Col  1  A: #
        ws.cell(next_row, 1,  value=row_num)
        # Col  2  B: Date Run
        ws.cell(next_row, 2,  value=today_str)
        # Col  3  C: Source (badge styled below)
        ws.cell(next_row, 3,  value=source_label)
        # Col  4  D: Company
        ws.cell(next_row, 4,  value=job.get("company", ""))
        # Col  5  E: Job Title
        ws.cell(next_row, 5,  value=job.get("title", ""))
        # Col  6  F: Job Level
        ws.cell(next_row, 6,  value=job.get("job_level", ""))
        # Col  7  G: Industry
        ws.cell(next_row, 7,  value=job.get("company_industry", ""))
        # Col  8  H: Location
        ws.cell(next_row, 8,  value=job.get("location", "Remote"))
        # Col  9  I: Job Type
        ws.cell(next_row, 9,  value=job.get("job_type", "Full-time"))
        # Col 10  J: Date Posted
        ws.cell(next_row, 10, value=job.get("date_posted", today_str))
        # Col 11  K: Days Since Posted — live Excel formula
        ws.cell(next_row, 11, value=f'=IF(J{next_row}="","",TODAY()-DATEVALUE(J{next_row}))')
        # Col 12  L: Salary Range
        ws.cell(next_row, 12, value=_format_salary(job))
        # Col 13  M: Job URL
        ws.cell(next_row, 13, value=job.get("job_url", ""))
        # Col 14  N: Company URL
        ws.cell(next_row, 14, value=job.get("company_url", ""))
        # Col 15  O: Recruiter Email
        ws.cell(next_row, 15, value=job.get("recruiter_emails", ""))
        # Col 16  P: Resume File
        ws.cell(next_row, 16, value=resume_name)
        # Col 17  Q: Status (default "To Apply" — user updates manually)
        ws.cell(next_row, 17, value="To Apply")
        # Col 18  R: Notes (blank — user fills)
        ws.cell(next_row, 18, value="")

        _style_row(ws, next_row, source_label, "To Apply", alt)
        ws.row_dimensions[next_row].height = 20
        rows_added += 1

        logger.info(
            f"  📊 [{source_label}] {job.get('title')} @ {job.get('company')} → {resume_name}"
        )

    # Update subtitle timestamp
    ws["A2"].value = (
        f"Auto-updated daily · Sites: LinkedIn, Indeed, ZipRecruiter · "
        f"Worldwide Remote · Last refresh: {datetime.today().strftime('%B %d, %Y')}"
    )

    wb.save(tracker_path)
    logger.info(f"✅ Tracker updated: {Path(tracker_path).resolve()} (+{rows_added} rows)")
    return rows_added
